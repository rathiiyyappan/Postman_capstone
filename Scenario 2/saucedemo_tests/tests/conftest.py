import pytest
from openpyxl import load_workbook

@pytest.fixture
def read_excel_data(request):
    data_file = request.config.getoption("--data-file")
    workbook = load_workbook(f"data/{data_file}")
    sheet = workbook.active
    row = request.node.get_closest_marker("excel_data").args[0]
    return sheet.cell(row=row, column=1).value, sheet.cell(row=row, column=2).value, sheet.cell(row=row, column=3).value

def pytest_addoption(parser):
    parser.addoption("--data-file", action="store", default="testdata.xlsx")
